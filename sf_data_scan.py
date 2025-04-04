import json
import requests
import time
import openpyxl
from snowflake.sqlalchemy import URL
from sqlalchemy import create_engine, text


# Load Snowflake credentials from file
def load_credentials(file_path):
    with open(file_path, 'r') as file:
        return json.load(file)


# Read table list from file
def load_table_list(file_path):
    with open(file_path, 'r') as file:
        return [line.strip() for line in file.readlines() if line.strip()]


# Establish Snowflake connection
def get_snowflake_connection(credentials):
    engine = create_engine(
        URL(
            account=credentials['account'],
            user=credentials['user'],
            password=credentials['password'],
            warehouse=credentials['warehouse'],
            role=credentials['role'],
            application="protecto"
        ),
        connect_args={'client_session_keep_alive': True}
    )
    return engine


# Fetch data from Snowflake
def fetch_data_from_snowflake(engine, table_name, limit, offset):
    query = f"SELECT * FROM {table_name} LIMIT {limit} OFFSET {offset}"
    with engine.connect() as connection:
        result = connection.execute(text(query))
        rows = result.fetchall()
        columns = list(result.keys())
    return columns, rows


# Split columns into chunks of 5
def split_columns(columns, rows, chunk_size=5):
    for i in range(0, len(columns), chunk_size):
        yield columns[i:i + chunk_size], [[row[j] for j in range(i, min(i + chunk_size, len(columns)))] for row in rows]


# Prepare API payload
def prepare_payload(data_source_name, object_name, columns, rows):
    payload = {
        "data_source_name": data_source_name,
        "object_name": object_name.split('.'),
        "data_samples": []
    }

    column_samples = {col: [] for col in columns}

    for row in rows:
        for col, value in zip(columns, row):
            column_samples[col].append(f"{col}: {value}")

    for column_name, samples in column_samples.items():
        payload["data_samples"].append({
            "column_name": column_name,
            "samples": samples
        })

    return payload


# Submit data scan
def submit_data_scan(base_url, auth_key, payload):
    url = f'{base_url}/data-scan/data-scan-async'
    headers = {'Authorization': f'Bearer {auth_key}', 'Content-Type': 'application/json'}
    response = requests.put(url, headers=headers, data=json.dumps([payload]))
    response_data = response.json()
    if response_data.get('success'):
        return response_data['data']['tracking_id']
    else:
        raise Exception("Error in submitting data scan: " + str(response_data.get('error', {}).get('message')))


# Track scan status
def track_status(base_url, auth_key, tracking_id):
    url = f'{base_url}/data-scan/status'
    headers = {'Authorization': f'Bearer {auth_key}', 'Content-Type': 'application/json'}
    while True:
        response = requests.put(url, headers=headers, data=json.dumps([tracking_id]))
        response_data = response.json()
        if response_data.get('success'):
            status = response_data['data'][0]['request_status']
            print(f"Tracking ID {tracking_id}: {status}")
            if status == 'SUCCESS':
                return
            elif status == 'ERROR':
                raise Exception("Data scan failed: " + response_data.get('error', {}).get('message'))
        time.sleep(5)


# Get report details with pagination
def get_report(base_url, auth_key, data_source_name, object_name):
    url = f'{base_url}/data-scan/objects/details'
    headers = {'Authorization': f'Bearer {auth_key}', 'Content-Type': 'application/json'}

    all_data = {"details": []}
    next_page_token = None

    while next_page_token is not None or not all_data["details"]:
        payload = {
            "data": {"data_source_name": data_source_name, "object_name": object_name.split('.')},
            "next_page_token": next_page_token
        }
        response = requests.put(url, headers=headers, data=json.dumps(payload))
        response_data = response.json()

        if response_data.get('success'):
            all_data["details"].extend(response_data['data']['details'])
            next_page_token = response_data.get('next_page_token')
        else:
            raise Exception("Error in getting report: " + response_data.get('error', {}).get('message'))

    return all_data

# Save report to Excel
def save_report(report, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['table_name','column_name', 'values_scanned', 'identified_count', 'identified_percentage', 'pi_type'])
    for detail in report['details']:
        for column in detail['columns']:
            pi_details = column.get('ml_identified_pi_details') or []
            first_row = ws.max_row + 1
            if pi_details:
                for pi_detail in pi_details:
                    ws.append(['.'.join(detail["object_name"]),column['column_name'], column['values_scanned'], pi_detail['identified_count'],
                               pi_detail['identified_percentage'], pi_detail['pi_type']])
                last_row = ws.max_row
                if last_row > first_row:
                    ws.merge_cells(start_row=first_row, start_column=1, end_row=last_row, end_column=1)
                    ws.merge_cells(start_row=first_row, start_column=2, end_row=last_row, end_column=2)
            else:
                ws.append(['.'.join(detail["object_name"]),column['column_name'], column['values_scanned'], 0, 0, None])
    wb.save(output_file)


# Main function
def main(base_url, credentials_file, table_list_file, num_rows, output_file):
    data_source_name = "SF_DS"
    credentials = load_credentials(credentials_file)
    auth_key = credentials["protecto_api_key"]
    table_list = load_table_list(table_list_file)
    engine = get_snowflake_connection(credentials)
    no_rows_per_call = 50
    report = {"details":[]}
    for table in table_list:
        print(f"Processing table: {table}")
        for i in range(0,num_rows,no_rows_per_call):
            columns, rows = fetch_data_from_snowflake(engine, table, no_rows_per_call, i)
            tracking_ids = []
            for chunk_columns, chunk_rows in split_columns(columns, rows):
                payload = prepare_payload(data_source_name, table, chunk_columns, chunk_rows)
                tracking_id = submit_data_scan(base_url, auth_key, payload)
                tracking_ids.append(tracking_id)

            for tracking_id in tracking_ids:
                track_status(base_url, auth_key, tracking_id)
                print(f"Tracking ID {tracking_id} completed successfully.")

        resp = get_report(base_url, auth_key, data_source_name, table)
        report["details"].extend(resp["details"])
    save_report(report, output_file)
    print("All tables processed successfully.")


# Execute script
if __name__ == '__main__':
    BASE_URL = "https://protecto-trial.protecto.ai/api/vault"
    CREDENTIALS_FILE = "credentials.json"
    TABLE_LIST_FILE = "tables.txt"
    NUM_ROWS = 100  # Adjust as needed
    OUTPUT_FILE = "data_scan_report.xlsx"

    main(BASE_URL, CREDENTIALS_FILE, TABLE_LIST_FILE, NUM_ROWS, OUTPUT_FILE)
